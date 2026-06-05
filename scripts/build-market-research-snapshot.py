"""
市場調査タブ用 Excel → JSON スナップショット変換スクリプト

入力: documents/ver2.2000~2025年各自治体出生数・保育所定員・利用人数.xlsx
出力: public/snapshots/market-research/data.json

実行: C:/Python314/python.exe scripts/build-market-research-snapshot.py
"""

from __future__ import annotations

import json
import sys
from datetime import datetime, timezone, timedelta
from pathlib import Path

import openpyxl

# ===== 全国地方公共団体コード（5桁）マッピング =====
# 総務省「全国地方公共団体コード」https://www.soumu.go.jp/denshijiti/code.html
CITY_CODE_MAP: dict[tuple[str, str], str] = {
    # 埼玉県 (11)
    ("埼玉県", "蕨市"): "11225",
    ("埼玉県", "川口市"): "11203",
    ("埼玉県", "和光市"): "11229",
    ("埼玉県", "朝霞市"): "11227",
    ("埼玉県", "新座市"): "11230",
    ("埼玉県", "上尾市"): "11219",
    ("埼玉県", "白岡市"): "11246",
    ("埼玉県", "春日部市"): "11214",
    ("埼玉県", "さいたま市"): "11100",
    ("埼玉県", "さいたま市西区"): "11101",
    ("埼玉県", "さいたま市南区"): "11108",
    ("埼玉県", "さいたま市浦和区"): "11107",
    # 千葉県 (12)
    ("千葉県", "千葉市"): "12100",
    ("千葉県", "市川市"): "12203",
    ("千葉県", "柏市"): "12217",
    ("千葉県", "成田市"): "12211",
    ("千葉県", "習志野市"): "12216",
    ("千葉県", "四街道市"): "12228",
    ("千葉県", "富里市"): "12233",
    ("千葉県", "東金市"): "12213",
    ("千葉県", "八千代市"): "12221",
    ("千葉県", "佐倉市"): "12212",
    # 東京都 (13)
    ("東京都", "大田区"): "13111",
    ("東京都", "練馬区"): "13120",
    ("東京都", "目黒区"): "13110",
    ("東京都", "杉並区"): "13115",
    ("東京都", "中野区"): "13114",
    ("東京都", "稲城市"): "13225",
    ("東京都", "港区"): "13103",
    ("東京都", "品川区"): "13109",
    # 神奈川県 (14)
    ("神奈川県", "相模原市"): "14150",
    ("神奈川県", "大和市"): "14213",
    ("神奈川県", "座間市"): "14216",
    ("神奈川県", "川崎市中原区"): "14133",
    # 茨城県 (08) - 1都3県外だが既存出店圏
    ("茨城県", "土浦市"): "08203",
    ("茨城県", "古河市"): "08204",
    ("茨城県", "つくば市"): "08220",
    # 三重県 (24) - 1都3県外だが既存出店圏
    ("三重県", "伊賀市"): "24216",
}

PREF_CODE_MAP = {
    "茨城県": "08",
    "埼玉県": "11",
    "千葉県": "12",
    "東京都": "13",
    "神奈川県": "14",
    "三重県": "24",
}

# 首都圏1都3県の県コード
METROPOLITAN_PREF_CODES = {"11", "12", "13", "14"}

# データテーブル内の各セクションの開始行（ヘッダ行）と metric キー
# row+1 が「都道府県/市区町村」の列見出し行、row+2 から実データ
SECTION_DEFS = [
    {"start_row": 1, "header_row": 2, "data_start": 3, "data_end": 37, "metric": "births", "label": "出生数"},
    {"start_row": 115, "header_row": 116, "data_start": 117, "data_end": 155, "metric": "capacity", "label": "保育所定員数(認可+小規模)"},
    {"start_row": 156, "header_row": 157, "data_start": 158, "data_end": 196, "metric": "enrollment", "label": "保育所利用人数(認可+小規模)"},
    {"start_row": 197, "header_row": 198, "data_start": 199, "data_end": 236, "metric": "utilization", "label": "保育所充足率(%)"},
]

YEARS = list(range(2000, 2026))  # 2000-2025年度


def to_number(v):
    """セル値を数値化、None / 文字列 / エラーは None を返す"""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v) if isinstance(v, float) else v
    try:
        return float(str(v).replace(",", ""))
    except (ValueError, AttributeError):
        return None


def main():
    repo_root = Path(__file__).resolve().parents[1]
    excel_path = repo_root.parent.parent / "documents" / "ver2.2000~2025年各自治体出生数・保育所定員・利用人数.xlsx"
    out_path = repo_root / "public" / "snapshots" / "market-research" / "data.json"
    out_path.parent.mkdir(parents=True, exist_ok=True)

    if not excel_path.exists():
        sys.stderr.write(f"ERROR: Excel not found at {excel_path}\n")
        sys.exit(1)

    print(f"Loading: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["データテーブル"]

    # cities: code -> { code, prefCode, prefName, name, isMetropolitan, data: { births: [], capacity: [], ... } }
    cities: dict[str, dict] = {}
    unmapped: list[tuple[str, str]] = []

    for section in SECTION_DEFS:
        metric = section["metric"]
        # Validate year columns: header_row col 3 = 2000年度, col 28 = 2025年度
        # Build column → year map (col index starts at 1, A=都道府県, B=市区町村, C=2000年度...)
        header_row = section["header_row"]
        col_to_year: dict[int, int] = {}
        for col in range(3, 3 + len(YEARS)):
            year = YEARS[col - 3]
            col_to_year[col] = year

        for r in range(section["data_start"], section["data_end"] + 1):
            pref = ws.cell(r, 1).value
            city = ws.cell(r, 2).value
            if not pref or not city:
                continue
            key = (pref.strip(), city.strip())
            code = CITY_CODE_MAP.get(key)
            if not code:
                if key not in unmapped:
                    unmapped.append(key)
                continue

            if code not in cities:
                pref_code = PREF_CODE_MAP[key[0]]
                cities[code] = {
                    "code": code,
                    "prefCode": pref_code,
                    "prefName": key[0],
                    "name": key[1],
                    "isMetropolitan": pref_code in METROPOLITAN_PREF_CODES,
                    "data": {"births": [], "capacity": [], "enrollment": [], "utilization": []},
                }

            series = []
            for col, year in col_to_year.items():
                v = to_number(ws.cell(r, col).value)
                series.append(v)
            cities[code]["data"][metric] = series

    if unmapped:
        sys.stderr.write("WARNING: 未マッピングの市区町村:\n")
        for p, c in unmapped:
            sys.stderr.write(f"  - {p} | {c}\n")

    # 各 city の data に欠けている metric は同じ長さの None 埋め
    for code, city in cities.items():
        for metric in ["births", "capacity", "enrollment", "utilization"]:
            if not city["data"].get(metric):
                city["data"][metric] = [None] * len(YEARS)

    jst = timezone(timedelta(hours=9))
    generated_at = datetime.now(jst).isoformat()

    output = {
        "meta": {
            "generatedAt": generated_at,
            "source": "社長共有 ver2.2000~2025年各自治体出生数・保育所定員・利用人数.xlsx",
            "years": YEARS,
            "metrics": {
                "births": "出生数 (人/年度)",
                "capacity": "保育所定員数 (認可+小規模, 4/1時点)",
                "enrollment": "保育所利用人数 (認可+小規模, 4/1時点)",
                "utilization": "保育所充足率 (%, 利用÷定員)",
            },
            "metropolitanPrefCodes": sorted(METROPOLITAN_PREF_CODES),
            "totalCities": len(cities),
            "metropolitanCities": sum(1 for c in cities.values() if c["isMetropolitan"]),
        },
        "cities": cities,
    }

    out_path.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Written: {out_path}")
    print(f"  Total cities: {len(cities)}")
    print(f"  Metropolitan (1都3県): {output['meta']['metropolitanCities']}")
    print(f"  Years: {YEARS[0]}-{YEARS[-1]}")


if __name__ == "__main__":
    main()
